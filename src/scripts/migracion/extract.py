# -*- coding: utf-8 -*-
"""
Etapa 1 de la carga inicial — Extractor del Excel del cliente.

Lee la planilla operativa (COBRANZA ARTICULO ...xlsx) y genera en `migracion/`
(carpeta GITIGNORED, contiene datos personales):

  - datos.json             → estado normalizado listo para migrate.load.js
  - clientes_revision.csv  → deduplicación propuesta de clientes (revisa el cliente)
  - excluidos_usd.csv      → operaciones en dólares excluidas (confirma el cliente)
  - anomalias.csv          → filas que requieren decisión humana
  - dni_asignados.csv      → mapeo cliente/usuario → DNI sintético
  - tasas_a_definir.csv    → combos frecuencia×cuotas reales para que el cliente
                             complete las tasas de operaciones nuevas

NO toca ninguna base de datos. Ver docs/plan-carga-inicial-produccion.md.

Uso:
    python src/scripts/migracion/extract.py "C:/ruta/COBRANZA ARTICULO 07-07-2026.xlsx"
"""
import csv
import json
import math
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl. Instalar con: pip install openpyxl")

# Consola Windows (cp1252) no soporta todos los caracteres del reporte.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ── Configuración ────────────────────────────────────────────────────────────
DNI_CLIENTES_DESDE = 99_000_001  # fuera del rango de DNIs argentinos reales
DNI_USUARIOS_DESDE = 99_999_001
UMBRAL_USD = 5_000               # cuota < $5.000 → operación en USD sin etiquetar
MARCA_MIGRACION = "[MIGRACION 2026-07]"
# Un usuario operativo por ZONA DE COBRANZA real del Excel (se derivan de los
# créditos, no de una lista fija). Las planillas solo se generan para
# COLLECTOR/SELLER_COLLECTOR (collections.service valida el rol), por eso
# Samuel — que además es ADMIN en producción — necesita esta cuenta operativa
# SEPARADA de la de admin (su DNI real ya lo usa el admin → DNI sintético).
# Tadeo y Leandro NO son zonas de cobranza → no se crean acá; sus cuentas
# ADMIN de producción no se tocan.
ROL_OPERATIVO = "SELLER_COLLECTOR"  # definido por el cliente: rol mixto para todos

DIAS_SEMANA = {
    "lunes": 1, "martes": 2, "miercoles": 3, "miércoles": 3,
    "jueves": 4, "viernes": 5, "sabado": 6, "sábado": 6, "domingo": 0,
}

SALIDA = Path(__file__).resolve().parents[3] / "migracion"


# ── Helpers ──────────────────────────────────────────────────────────────────
def normalizar(texto):
    """Mayúsculas, sin tildes, espacios colapsados."""
    if texto is None:
        return ""
    s = str(texto).strip().upper()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s)


def nombre_base(nombre_norm):
    """Quita sufijo numérico de créditos repetidos: 'BELEN QUIROGA 2' → 'BELEN QUIROGA'."""
    return re.sub(r"\s+\d{1,2}$", "", nombre_norm)


def _lev(a, b):
    """Distancia de edición (Levenshtein)."""
    if a == b:
        return 0
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        cur = [i]
        for j, cb in enumerate(b, 1):
            cur.append(min(prev[j] + 1, cur[j - 1] + 1, prev[j - 1] + (ca != cb)))
        prev = cur
    return prev[-1]


def nombres_similares(a, b):
    """True si dos nombres normalizados son la misma persona con typos: cada
    token de la parte común coincide o difiere por pocos caracteres. Rescata
    'jesica torino' vs 'jesica torno' pero rechaza 'maxi rubio' vs 'maxi matana'
    (apellidos distintos)."""
    for x, y in zip(a.split(), b.split()):
        d = _lev(x, y)
        if d and not (d <= 2 and d <= 0.4 * max(len(x), len(y))):
            return False
    return True


def a_titulo(texto):
    """'CARO QUIRINO' → 'Caro Quirino' (nombre presentable)."""
    return " ".join(w.capitalize() for w in normalizar(texto).split())


def num(valor):
    """Convierte celda a float; None si no es numérica."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    try:
        return float(str(valor).replace(",", ".").strip())
    except ValueError:
        return None


def ctas_de_celda(valor):
    """CTAS puede venir como número o como fecha-artefacto de Excel
    (ej. datetime(1900,1,24) = serial 24 → 24 cuotas). Para seriales < 60 la
    época correcta es 1899-12-31 (bug del 29/02/1900 de Excel)."""
    if isinstance(valor, datetime):
        epoca = datetime(1899, 12, 31) if valor < datetime(1900, 3, 1) else datetime(1899, 12, 30)
        return (valor - epoca).days
    n = num(valor)
    return int(n) if n is not None else None


def fecha_iso(valor):
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    return None


# ── Capital desde la Observación de CONTROL DE PAGOS ─────────────────────────
# El capital viene en taquigrafía: "300MIL"=300.000, "1,370MILL"=1.370.000, y
# a veces "1,5MIL" son millones. Se desambigua la magnitud con el Total (que YA
# incluye interés): el capital financiado tiene que caer entre ~35% y ~97% del
# Total, así se descarta la interpretación equivocada (regla del cliente).
BANDA_CAPITAL = (0.35, 0.97)


def parse_capital(obs, total):
    """Lee el capital de la Observación y desambigua mil/millón contra el Total.
    @returns (capital|None, flag) con flag ∈ ok|ambiguo|fuera_banda|sin_numero|
    sin_obs|sin_total|sin_ficha — para el CSV de revisión."""
    if obs is None or str(obs).strip() == "":
        return None, "sin_obs"
    if isinstance(obs, (int, float)):
        cands = [float(obs)]
    else:
        s = normalizar(obs)
        m = re.search(r"(\d+(?:[.,]\d+)?)\s*(MILL|MIL)?", s.replace("REF", ""))
        if not m:
            return None, "sin_numero"
        base = float(m.group(1).replace(".", "").replace(",", "."))
        mag = m.group(2)
        if mag == "MILL":
            cands = [base * 1_000_000, base * 1_000]
        elif mag == "MIL":
            cands = [base * 1_000, base * 1_000_000]
        else:
            cands = [base * 1_000, base * 1_000_000, base]
    if not total:
        return None, "sin_total"
    lo, hi = BANDA_CAPITAL
    inband = [c for c in cands if lo * total <= c <= hi * total]
    if not inband:
        return None, "fuera_banda"
    return round(inband[0], 2), ("ambiguo" if len(inband) > 1 else "ok")


# ── Lectura del maestro (Planilla COBRANZA) ──────────────────────────────────
def leer_maestro(wb):
    ws = wb["Planilla COBRANZA"]
    creditos, excluidos, anomalias = [], [], []
    seccion = "SEMANAL"  # antes del encabezado MENSUALES

    for i, row in enumerate(
        ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=12, values_only=True), 3
    ):
        nombre = row[0]
        if nombre is None or str(nombre).strip() == "":
            continue

        ctas, imp = ctas_de_celda(row[4]), num(row[6])

        # encabezado de sección: fila con nombre pero sin CTAS ni IMP
        if ctas is None and imp is None:
            if "MENSUAL" in normalizar(nombre):
                seccion = "MENSUAL"
            else:
                anomalias.append((i, str(nombre), "fila sin CTAS/IMP no reconocida"))
            continue

        zona = normalizar(row[1])
        producto = str(row[2]).strip() if row[2] is not None else ""
        pagado = num(row[8]) or 0.0
        saldo = num(row[9]) or 0.0
        plan = row[10]

        base = {
            "fila_excel": i,
            "nombre": str(nombre).strip(),
            "zona": zona,
            "producto": producto,
            "ctas": ctas,
            "imp": imp,
            "pagado": pagado,
            "saldo": saldo,
        }

        # ── exclusión USD: etiqueta explícita o monto imposible en pesos ──
        if "USD" in producto.upper() or (imp is not None and imp < UMBRAL_USD):
            excluidos.append(base)
            continue

        # ── validación de consistencia ──
        if ctas is None or imp is None:
            anomalias.append((i, base["nombre"], f"CTAS o IMP ilegible (ctas={row[4]!r}, imp={row[6]!r})"))
            continue
        total_teorico = ctas * imp
        if abs(total_teorico - (pagado + saldo)) > 1:
            anomalias.append(
                (i, base["nombre"],
                 f"CTASxIMP={total_teorico:.0f} != PAGADO+SALDO={(pagado + saldo):.0f}")
            )
            continue

        # ── saldados: quedan fuera (decisión §6.5: no cargar) ──
        if saldo == 0:
            anomalias.append((i, base["nombre"], "SALDO=0 (ya saldado — no se carga, decisión §6.5)"))
            continue

        # ── frecuencia según PLAN + sección ──
        plan_norm = normalizar(plan) if not isinstance(plan, datetime) else None
        if isinstance(plan, datetime):
            frecuencia, dia_pago, proximo_venc = "MONTHLY", None, fecha_iso(plan)
        elif plan_norm and plan_norm.lower() in (d.upper().lower() for d in DIAS_SEMANA):
            frecuencia, dia_pago, proximo_venc = "WEEKLY", str(plan).strip().lower(), None
        elif plan_norm and "DIAR" in plan_norm:
            frecuencia, dia_pago, proximo_venc = "DAILY", None, None
        elif plan_norm and "QUINCEN" in plan_norm:
            frecuencia, dia_pago, proximo_venc = "BIWEEKLY", None, None
        else:
            anomalias.append((i, base["nombre"], f"PLAN ilegible: {plan!r} (¿día de pago?) — bloqueante"))
            continue

        # coherencia sección/frecuencia (informativo)
        if seccion == "MENSUAL" and frecuencia == "WEEKLY":
            anomalias.append((i, base["nombre"], "día de semana en sección MENSUALES — revisar"))
            continue

        # ── estado de cuotas ──
        cuotas_pagadas = math.floor(round(pagado / imp, 6))
        resto_parcial = round(pagado - cuotas_pagadas * imp, 2)

        base.update({
            "tipo": "LOAN" if producto.upper().startswith("PRESTAMO") else "SALE",
            "frecuencia": frecuencia,
            "dia_pago": dia_pago,            # WEEKLY: lunes..sabado
            "proximo_venc": proximo_venc,    # MONTHLY: fecha (puede ser pasada = mora)
            "cuotas_pagadas": cuotas_pagadas,
            "resto_parcial": resto_parcial,  # >0 → una cuota PARTIAL
        })
        creditos.append(base)

    return creditos, excluidos, anomalias


# ── Lectura de fichas (CONTROL DE PAGOS) — mejor esfuerzo ────────────────────
def leer_fichas(wb):
    """Devuelve fichas: nombre_norm → lista de dicts con f_inicio y cantidad de pagos."""
    ws = wb["CONTROL DE PAGOS"]
    fichas = []
    actual = None
    esperando_fechas = False

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=13, values_only=True):
        etiqueta = normalizar(row[0]) if row[0] is not None else ""

        if etiqueta == "NOMBRE:":
            if actual:
                fichas.append(actual)
            actual = {"nombre": str(row[1]).strip() if row[1] is not None else "",
                      "producto": "", "obs": None, "f_entrega": None, "f_inicio": None,
                      "total": None, "saldo": None, "pagos": 0, "pagado_suma": 0.0}
            esperando_fechas = False
        elif actual is None:
            continue
        elif etiqueta == "PRODUCTO:":
            actual["producto"] = str(row[1]).strip() if row[1] is not None else ""
        elif etiqueta == "OBSERVACION:":
            actual["obs"] = row[1]  # capital en taquigrafía ("300MIL", etc.)
        elif etiqueta == "F.ENTREGA":
            esperando_fechas = True
        elif esperando_fechas:
            actual["f_entrega"] = fecha_iso(row[0])
            actual["f_inicio"] = fecha_iso(row[1])
            esperando_fechas = False
        elif etiqueta == "TOTAL:":
            actual["total"] = num(row[1])
            actual["saldo"] = num(row[3])

        # entradas de pago en 3 grupos de columnas (N°, monto, fecha)
        if actual:
            for b in (4, 7, 10):
                monto = row[b + 1] if len(row) > b + 1 else None
                fch = row[b + 2] if len(row) > b + 2 else None
                if isinstance(monto, (int, float)) and fch is not None:
                    actual["pagos"] += 1
                    actual["pagado_suma"] += float(monto)

    if actual:
        fichas.append(actual)

    indice = defaultdict(list)
    for f in fichas:
        indice[nombre_base(normalizar(f["nombre"]))].append(f)
    return fichas, indice


def cruzar_fichas(creditos, indice_fichas, fichas):
    """Enriquece cada crédito con la ficha de CONTROL DE PAGOS que matchea:
      - por nombre (único candidato, o desempate/valor compartido por saldo);
      - si el nombre no matchea (typos entre las dos hojas, ej. "jesica torino"
        vs "jesica torno"), FALLBACK por SALDO exacto + primer token del nombre.
    De la ficha toma fechas F.INICIO/F.ENTREGA (ancla de la 1ª cuota), capital
    (de Observación, desambiguado con el total) y la tasa derivada. Sin ficha →
    capital = total, interés 0, y el generador ancla a la fecha de carga."""
    # índice por saldo redondeado para el fallback por monto
    por_saldo_idx = defaultdict(list)
    for f in fichas:
        if f.get("saldo") is not None:
            por_saldo_idx[round(f["saldo"])].append(f)

    con_fecha = con_capital = 0
    for c in creditos:
        candidatos = indice_fichas.get(nombre_base(normalizar(c["nombre"])), [])
        por_saldo = [f for f in candidatos if f["saldo"] is not None
                     and abs(f["saldo"] - c["saldo"]) < 1]
        elegido = None
        if len(candidatos) == 1:
            elegido = candidatos[0]
        elif len(por_saldo) == 1:
            elegido = por_saldo[0]

        # Fuente de fecha/capital: la ficha elegida, o —si quedó ambiguo (mismo
        # nombre y saldo, ej. "cuni 1"/"cuni 2")— el valor COMPARTIDO por los
        # candidatos: si todos coinciden en F.ENTREGA/capital, da igual cuál es
        # cuál y se puede tomar igual.
        fuente = [elegido] if elegido else (por_saldo or candidatos)

        # Fallback por SALDO exacto + nombre similar: rescata créditos cuyo
        # nombre difiere entre las dos hojas por un typo (jesica torino vs torno).
        # Exige saldo idéntico, mismo primer nombre y nombres similares (rechaza
        # mismo primer nombre con apellido distinto, ej. maxi rubio / maxi matana).
        if not fuente:
            base_c = nombre_base(normalizar(c["nombre"]))
            tok = base_c.split(" ")[0]
            cs = [f for f in por_saldo_idx.get(round(c["saldo"]), [])
                  if abs(f["saldo"] - c["saldo"]) < 1
                  and normalizar(f["nombre"]).split(" ")[:1] == [tok]
                  and nombres_similares(base_c, nombre_base(normalizar(f["nombre"])))]
            if len(cs) == 1:
                elegido = cs[0]
                fuente = cs
                c["match_typo"] = True

        def compartido(campo):
            vals = {f.get(campo) for f in fuente if f.get(campo)}
            return next(iter(vals)) if len(vals) == 1 else None

        total_maestro = round(c["ctas"] * c["imp"], 2)  # total a devolver (con interés)
        c["f_inicio"] = compartido("f_inicio")
        c["f_entrega"] = compartido("f_entrega")
        c["pagos_historicos"] = elegido["pagos"] if elegido else None
        capital, flag = parse_capital(compartido("obs"), total_maestro)
        if not fuente:
            flag = "sin_ficha"

        # ancla del cronograma: F.INICIO (1ª cuota) o, si falta, F.ENTREGA
        c["ancla_fecha"] = c["f_inicio"] or c["f_entrega"]
        c["ancla_fuente"] = "F.INICIO" if c["f_inicio"] else ("F.ENTREGA" if c["f_entrega"] else None)
        # capital financiado (→ total_amount) y tasa congelada (→ interest_rate)
        c["capital"] = capital
        c["capital_flag"] = flag
        c["total_amount"] = capital if capital else total_maestro
        c["interest_rate"] = round(total_maestro / capital - 1, 4) if capital else 0.0

        if c["ancla_fecha"]:
            con_fecha += 1
        if capital:
            con_capital += 1
    return con_fecha, con_capital


# ── Catálogo (Hoja1) ─────────────────────────────────────────────────────────
def leer_catalogo(wb):
    """Catálogo de artículos (Hoja1). Localiza el encabezado ARTICULOS de forma
    DINÁMICA: entre snapshots el cliente movió las columnas de esta hoja
    (14-07 en col 7, 28-07 en col 8), así que anclar a una columna fija hacía
    que se leyeran 0 productos silenciosamente. A partir del encabezado:
    nombre | PRECIO/U (+1) | CANTIDAD (+2) | PRECIO TADEO (+3) |
    PRECIO LEANDRO (+4, = PRECIO FINAL → precio de lista §6.7)."""
    ws = wb["Hoja1"]
    col_art = fila_header = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, max_col=25, values_only=True), 1):
        for j, cell in enumerate(row):
            if normalizar(cell) == "ARTICULOS":  # primera ocurrencia = tabla principal
                col_art, fila_header = j, i      # j: 0-based → columna 1-based = j+1
                break
        if col_art is not None:
            break
    if col_art is None:
        print("   ⚠ Hoja1: no se encontró el encabezado 'ARTICULOS' — 0 productos")
        return []

    productos = []
    for row in ws.iter_rows(min_row=fila_header + 1, max_row=fila_header + 100,
                            min_col=col_art + 1, max_col=col_art + 5, values_only=True):
        nombre = row[0]
        if nombre is None or str(nombre).strip() == "":
            break  # fin del bloque contiguo de artículos
        costo, cantidad, precio_lista = num(row[1]), num(row[2]), num(row[4])
        productos.append({
            "nombre": a_titulo(nombre),
            "costo": round(costo) if costo else None,
            "stock": int(cantidad) if cantidad else 0,
            "precio_lista": round(precio_lista) if precio_lista else None,
        })
    return productos


# ── Armado de clientes + DNIs ────────────────────────────────────────────────
def armar_clientes(creditos):
    grupos = defaultdict(lambda: {"variantes": set(), "zonas": set(), "creditos": 0})
    for c in creditos:
        clave = nombre_base(normalizar(c["nombre"]))
        g = grupos[clave]
        g["variantes"].add(c["nombre"])
        g["zonas"].add(c["zona"])
        g["creditos"] += 1

    clientes, dni = [], DNI_CLIENTES_DESDE
    for clave in sorted(grupos):
        g = grupos[clave]
        clientes.append({
            "nombre": a_titulo(clave),
            "clave": clave,
            "dni": str(dni),
            "zona": sorted(g["zonas"])[0],
            "zonas": sorted(g["zonas"]),
            "variantes": sorted(g["variantes"]),
            "creditos": g["creditos"],
            "revisar": len(g["zonas"]) > 1,  # mismo nombre en 2 zonas: ¿2 personas?
        })
        dni += 1

    # referencia cliente en cada crédito
    dni_por_clave = {c["clave"]: c["dni"] for c in clientes}
    for c in creditos:
        c["cliente_dni"] = dni_por_clave[nombre_base(normalizar(c["nombre"]))]
    return clientes


def armar_usuarios(creditos):
    """Un usuario operativo por zona de cobranza presente en los créditos."""
    zonas = sorted({c["zona"] for c in creditos if c["zona"]})
    usuarios, dni = [], DNI_USUARIOS_DESDE
    for zona in zonas:
        usuarios.append({"nombre": a_titulo(zona), "usuario": zona.lower(),
                         "dni": str(dni), "rol": ROL_OPERATIVO})
        dni += 1
    return usuarios


# ── Salidas ──────────────────────────────────────────────────────────────────
def escribir_csv(ruta, encabezado, filas):
    with open(ruta, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(encabezado)
        w.writerows(filas)


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    origen = Path(sys.argv[1])
    if not origen.exists():
        sys.exit(f"No existe el archivo: {origen}")

    print(f"Leyendo {origen.name} ...")
    wb = openpyxl.load_workbook(origen, data_only=True, read_only=True)

    creditos, excluidos, anomalias = leer_maestro(wb)
    fichas, indice = leer_fichas(wb)
    con_fecha, con_capital = cruzar_fichas(creditos, indice, fichas)
    catalogo = leer_catalogo(wb)
    clientes = armar_clientes(creditos)
    usuarios = armar_usuarios(creditos)

    SALIDA.mkdir(exist_ok=True)

    # datos.json — insumo de migrate.load.js
    datos = {
        "meta": {
            "origen": origen.name,
            "generado": datetime.now().isoformat(timespec="seconds"),
            "marca": MARCA_MIGRACION,
            "totales": {
                "creditos": len(creditos),
                "clientes": len(clientes),
                "usuarios": len(usuarios),
                "productos_catalogo": len(catalogo),
                "excluidos_usd": len(excluidos),
                "anomalias": len(anomalias),
                "suma_saldo": round(sum(c["saldo"] for c in creditos), 2),
                "suma_capital": round(sum((c.get("capital") or c["total_amount"]) for c in creditos), 2),
                "creditos_con_fecha": con_fecha,
                "creditos_con_capital": con_capital,
            },
        },
        "usuarios": usuarios,
        "clientes": clientes,
        "productos": catalogo,
        "creditos": creditos,
    }
    with open(SALIDA / "datos.json", "w", encoding="utf-8") as f:
        json.dump(datos, f, ensure_ascii=False, indent=2)

    escribir_csv(SALIDA / "clientes_revision.csv",
                 ["nombre", "dni_asignado", "zonas", "creditos", "variantes_en_excel", "REVISAR"],
                 [[c["nombre"], c["dni"], " + ".join(c["zonas"]), c["creditos"],
                   " / ".join(c["variantes"]), "SI" if c["revisar"] else ""] for c in clientes])

    escribir_csv(SALIDA / "excluidos_usd.csv",
                 ["fila_excel", "nombre", "producto", "ctas", "imp", "saldo", "motivo"],
                 [[e["fila_excel"], e["nombre"], e["producto"], e["ctas"], e["imp"], e["saldo"],
                   "etiquetado USD" if "USD" in e["producto"].upper() else f"cuota < ${UMBRAL_USD} (USD sin etiquetar)"]
                  for e in excluidos])

    escribir_csv(SALIDA / "anomalias.csv",
                 ["fila_excel", "nombre", "detalle"],
                 [list(a) for a in anomalias])

    escribir_csv(SALIDA / "dni_asignados.csv",
                 ["tipo", "nombre", "dni_asignado"],
                 [["USUARIO", u["nombre"], u["dni"]] for u in usuarios] +
                 [["CLIENTE", c["nombre"], c["dni"]] for c in clientes])

    # Capital por crédito (leído de Observación). flag != ok/ambiguo → cae a
    # capital = total (interés 0): son los que conviene que el cliente complete.
    escribir_csv(SALIDA / "capital_revision.csv",
                 ["fila_excel", "nombre", "tipo", "ctas", "imp", "total_a_devolver",
                  "capital", "tasa_%", "flag"],
                 [[c["fila_excel"], c["nombre"], c["tipo"], c["ctas"], c["imp"],
                   round(c["ctas"] * c["imp"], 2), c.get("capital"),
                   round(c.get("interest_rate", 0) * 100, 1), c.get("capital_flag")]
                  for c in creditos])

    # Fecha ancla por crédito. fuente_fecha vacío → sin fecha real (el cronograma
    # cae al fallback desde la fecha de carga): a completar por el cliente.
    escribir_csv(SALIDA / "fechas_revision.csv",
                 ["fila_excel", "nombre", "frecuencia", "cuotas", "cuotas_pagadas",
                  "ancla_fecha", "fuente_fecha", "match_por_typo", "dia_pago", "proximo_venc"],
                 [[c["fila_excel"], c["nombre"], c["frecuencia"], c["ctas"],
                   c["cuotas_pagadas"], c.get("ancla_fecha"), c.get("ancla_fuente"),
                   "SI" if c.get("match_typo") else "", c.get("dia_pago"), c.get("proximo_venc")]
                  for c in creditos])

    # Separado por tipo: PRESTAMO → interest_rates (además necesita rango de
    # monto), VENTA → product_rates (además se define POR PRODUCTO). Los
    # créditos migrados no dependen de estas tasas; son para operar de ahí
    # en adelante.
    combos = Counter((c["tipo"], c["frecuencia"], c["ctas"]) for c in creditos)
    escribir_csv(SALIDA / "tasas_a_definir.csv",
                 ["destino", "frecuencia", "cuotas", "creditos_existentes",
                  "tasa_a_completar_por_cliente_%"],
                 [["PRESTAMOS (interest_rates)" if t == "LOAN" else "VENTAS (product_rates, por producto)",
                   f, ct, n, ""] for (t, f, ct), n in sorted(combos.items())])

    # ── Reporte ──
    t = datos["meta"]["totales"]
    freqs = Counter(c["frecuencia"] for c in creditos)
    tipos = Counter(c["tipo"] for c in creditos)
    print(f"""
── Extracción completa → {SALIDA} ──
  Créditos a cargar : {t['creditos']}  (LOAN {tipos.get('LOAN', 0)} / SALE {tipos.get('SALE', 0)})
  Por frecuencia    : {dict(freqs)}
  Clientes únicos   : {t['clientes']}  (DNIs desde {DNI_CLIENTES_DESDE})
  Usuarios          : {t['usuarios']}  (rol {ROL_OPERATIVO}, DNIs desde {DNI_USUARIOS_DESDE})
  Productos catálogo: {t['productos_catalogo']}
  Excluidos USD     : {t['excluidos_usd']}
  Anomalías         : {t['anomalias']}  ← revisar anomalias.csv
  Con fecha de ancla: {t['creditos_con_fecha']} / {t['creditos']}  (F.INICIO/F.ENTREGA → cronograma real; ver fechas_revision.csv)
  Con capital real  : {t['creditos_con_capital']} / {t['creditos']}  (resto: capital = total, interés 0; ver capital_revision.csv)
  SUMA SALDO        : ${t['suma_saldo']:,.0f}  ← verificar contra el sistema post-carga
  SUMA CAPITAL      : ${t['suma_capital']:,.0f}  (financiado, sin interés)
""")


if __name__ == "__main__":
    main()
